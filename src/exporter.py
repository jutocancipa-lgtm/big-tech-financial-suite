"""
exporter.py - Writes cleaned financial statements and ratios to an Excel file.

Uses openpyxl for formatting (bold headers, number format, column widths).
All paths are relative.
"""

from pathlib import Path
from typing import Optional

import pandas as pd


def _style_sheet(
    writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame
) -> None:
    """Apply basic formatting: bold index, auto column width, number format."""
    worksheet = writer.sheets[sheet_name]
    workbook = writer.book

    from openpyxl.styles import Font

    # Bold the header row
    header_cells = worksheet[1]
    for cell in header_cells:
        cell.font = Font(bold=True)

    # Number format for all data cells (millions with 2 decimals)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Auto-fit column widths (approximate)
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col)),
        )
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 4, 40)


def _write_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: Optional[pd.DataFrame],
    label: str,
) -> None:
    """Write a DataFrame to a sheet. If None, write a placeholder message."""
    actual_name = f"{label[:31]}"  # Excel limit: 31 chars
    if df is not None and not df.empty:
        df.to_excel(writer, sheet_name=actual_name)
        _style_sheet(writer, actual_name, df)
    else:
        placeholder = pd.DataFrame(
            {"Status": ["No data available for this statement."]}
        )
        placeholder.to_excel(writer, sheet_name=actual_name, index=False)


def export(
    ticker: str,
    income: Optional[pd.DataFrame],
    balance: Optional[pd.DataFrame],
    cashflow: Optional[pd.DataFrame],
    ratios: dict[str, Optional[pd.DataFrame]],
    data_dir: Path,
) -> Path:
    """Export all data for a single ticker into one Excel file.

    Returns the path to the created file.
    """
    data_dir.mkdir(parents=True, exist_ok=True)
    filepath = data_dir / f"{ticker}_financial_model.xlsx"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Core statements
        _write_sheet(writer, "Income Statement", income, "Income Statement")
        _write_sheet(writer, "Balance Sheet", balance, "Balance Sheet")
        _write_sheet(writer, "Cash Flow", cashflow, "Cash Flow Statement")

        # Ratio sheets
        for category, df in ratios.items():
            sheet = f"{category[:25]} Ratios"
            _write_sheet(writer, sheet, df, f"{category} Ratios")

    print(f"  Saved: {filepath}")
    return filepath
